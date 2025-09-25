import streamlit as st
import pandas as pd
import openpyxl
import os
from io import BytesIO
import tempfile

# Set page config
st.set_page_config(
    page_title="Price Editor",
    page_icon="💰",
    layout="wide"
)

def load_excel_file(file_path):
    """Load Excel file and return workbook"""
    try:
        return openpyxl.load_workbook(file_path)
    except Exception as e:
        st.error(f"Error loading Excel file: {e}")
        return None

def get_current_prices(workbook):
    """Get current values from D4-D7 in Prices sheet"""
    if 'Prices' not in workbook.sheetnames:
        return None
    
    prices_sheet = workbook['Prices']
    current_prices = {
        'D4': prices_sheet['D4'].value,
        'D5': prices_sheet['D5'].value,
        'D6': prices_sheet['D6'].value,
        'D7': prices_sheet['D7'].value
    }
    return current_prices

def update_prices(workbook, new_prices):
    """Update D4-D7 cells in Prices sheet and ensure calculations update"""
    if 'Prices' not in workbook.sheetnames:
        return False
    
    prices_sheet = workbook['Prices']
    prices_sheet['D4'] = new_prices['D4']
    prices_sheet['D5'] = new_prices['D5']
    prices_sheet['D6'] = new_prices['D6']
    prices_sheet['D7'] = new_prices['D7']
    
    # Set calculation mode to automatic to ensure formulas recalculate
    workbook.calculation.calcMode = 'auto'
    
    return True

def export_sheet_to_csv(workbook, sheet_name):
    """Export specified sheet to CSV bytes with proper formula calculation"""
    if sheet_name not in workbook.sheetnames:
        return None
    
    # Create a temporary file to save the workbook
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        try:
            # First save the workbook with updated values
            workbook.save(tmp_file.name)
            
            # Try multiple approaches to get calculated values
            try:
                # Method 1: Use xlwings if available (best for calculation)
                import xlwings as xw
                app = xw.App(visible=False)
                wb = app.books.open(tmp_file.name)
                wb.app.calculation = 'automatic'
                wb.save()
                wb.close()
                app.quit()
            except ImportError:
                # Method 2: Use openpyxl with data_only and manual calculation trigger
                pass
            
            # Load with openpyxl data_only=True to get values
            calc_wb = openpyxl.load_workbook(tmp_file.name, data_only=True)
            
            # If data_only still shows None, try to extract raw data and formulas
            if sheet_name in calc_wb.sheetnames:
                sheet = calc_wb[sheet_name]
                
                # Check if we're getting None values and try alternative approach
                has_none_values = False
                for row in sheet.iter_rows(min_row=1, max_row=min(10, sheet.max_row)):
                    for cell in row:
                        if cell.value is None and cell.coordinate in ['A1', 'B1', 'C1']:  # Check header area
                            continue
                        elif cell.value is None:
                            has_none_values = True
                            break
                    if has_none_values:
                        break
                
                if has_none_values:
                    # Fallback: load without data_only to preserve formulas, then try pandas with engine
                    calc_wb.close()
                    df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, engine='openpyxl')
                else:
                    # Save the calculated workbook
                    calc_wb.save(tmp_file.name)
                    calc_wb.close()
                    df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, engine='openpyxl')
            else:
                calc_wb.close()
                return None
            
            # Convert to CSV
            csv_string = df.to_csv(index=False)
            return csv_string.encode()
            
        except Exception as e:
            st.error(f"Export error: {e}")
            # Fallback: try basic export
            try:
                df = pd.read_excel(tmp_file.name, sheet_name=sheet_name)
                csv_string = df.to_csv(index=False)
                return csv_string.encode()
            except:
                return None
            
        finally:
            # Clean up temp file
            if os.path.exists(tmp_file.name):
                try:
                    os.unlink(tmp_file.name)
                except:
                    pass

def main():
    st.title("💰 Excel Price Editor")
    st.markdown("Edit prices in cells D4-D7 of the 'Prices' sheet and export the 'Export' sheet as CSV")
    
    # File handling
    excel_file_path = "your_workbook.xlsx"  # Default bundled file
    
    # Check if bundled file exists
    if os.path.exists(excel_file_path):
        st.success(f"✅ Using file: {excel_file_path}")
        use_bundled = True
    else:
        use_bundled = False
        st.warning("⚠️ Default file not found. Please upload an Excel file.")
    
    # File uploader (always show as backup option)
    uploaded_file = st.file_uploader(
        "📁 Upload Excel File (optional - overrides default)",
        type=['xlsx', 'xls'],
        help="Upload your Excel file containing 'Prices' and 'Export' sheets"
    )
    
    # Determine which file to use
    workbook = None
    temp_file_path = None
    
    if uploaded_file is not None:
        # Save uploaded file to temp location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_file_path = tmp_file.name
        
        workbook = load_excel_file(temp_file_path)
        st.info("📤 Using uploaded file")
        
    elif use_bundled:
        workbook = load_excel_file(excel_file_path)
    
    if workbook is None:
        st.stop()
    
    # Show available sheets
    st.sidebar.subheader("📋 Available Sheets")
    for sheet in workbook.sheetnames:
        icon = "💰" if sheet == "Prices" else "📊" if sheet == "Export" else "📄"
        st.sidebar.write(f"{icon} {sheet}")
    
    # Main interface
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("✏️ Edit Prices (Cells D4-D7)")
        
        # Get current prices
        current_prices = get_current_prices(workbook)
        
        if current_prices is None:
            st.error("❌ 'Prices' sheet not found in the workbook")
            st.stop()
        
        # Create form for editing prices
        with st.form("price_editor"):
            st.write("**Current Values:**")
            
            # Input fields for each price cell
            new_d4 = st.text_input(
                "D4 - Price 1:", 
                value=str(current_prices['D4']) if current_prices['D4'] is not None else "",
                help="Enter new value for cell D4"
            )
            
            new_d5 = st.text_input(
                "D5 - Price 2:", 
                value=str(current_prices['D5']) if current_prices['D5'] is not None else "",
                help="Enter new value for cell D5"
            )
            
            new_d6 = st.text_input(
                "D6 - Price 3:", 
                value=str(current_prices['D6']) if current_prices['D6'] is not None else "",
                help="Enter new value for cell D6"
            )
            
            new_d7 = st.text_input(
                "D7 - Price 4:", 
                value=str(current_prices['D7']) if current_prices['D7'] is not None else "",
                help="Enter new value for cell D7"
            )
            
            # Submit button
            submitted = st.form_submit_button("💾 Update Prices", type="primary")
            
            if submitted:
                # Prepare new prices (try to convert to float if possible, otherwise keep as string)
                new_prices = {}
                for cell, value in [('D4', new_d4), ('D5', new_d5), ('D6', new_d6), ('D7', new_d7)]:
                    try:
                        # Try to convert to float if it's a number
                        new_prices[cell] = float(value) if value and value.replace('.', '').replace('-', '').isdigit() else value
                    except:
                        new_prices[cell] = value
                
                # Update the workbook
                if update_prices(workbook, new_prices):
                    # Save the file
                    file_to_save = temp_file_path if temp_file_path else excel_file_path
                    try:
                        workbook.save(file_to_save)
                        st.success("✅ Prices updated successfully!")
                        
                        # Show updated values
                        st.write("**Updated Values:**")
                        for cell, value in new_prices.items():
                            st.write(f"• {cell}: {value}")
                        
                        st.info("💡 Export sheet will now reflect the updated prices!")
                        
                        # Auto-rerun to refresh the interface
                        st.rerun()
                        
                    except Exception as e:
                        st.error(f"❌ Error saving file: {e}")
                else:
                    st.error("❌ Failed to update prices")
    
    with col2:
        st.subheader("📊 Export Options")
        
        # Check if Export sheet exists
        if 'Export' in workbook.sheetnames:
            st.write("**Export Sheet Available** ✅")
            
            # Preview Export sheet (with updated values)
            with st.expander("👁️ Preview Export Sheet"):
                try:
                    # Create temp file to read with pandas
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                        workbook.save(tmp.name)
                        
                        # Try to read with calculated values
                        try:
                            preview_wb = openpyxl.load_workbook(tmp.name, data_only=True)
                            preview_wb.save(tmp.name)
                            preview_wb.close()
                            preview_df = pd.read_excel(tmp.name, sheet_name='Export')
                        except:
                            # Fallback to regular reading
                            preview_df = pd.read_excel(tmp.name, sheet_name='Export')
                        
                        st.dataframe(preview_df.head(10), use_container_width=True)
                        
                        # Show info about Variant Price column if it exists
                        if 'Variant Price' in preview_df.columns:
                            variant_price_col = preview_df['Variant Price']
                            non_null_count = variant_price_col.count()
                            st.write(f"**Variant Price column**: {non_null_count} calculated values out of {len(variant_price_col)} rows")
                            
                            if non_null_count == 0:
                                st.warning("⚠️ Variant Price column is empty. This might indicate formula calculation issues.")
                                st.write("**Troubleshooting tips:**")
                                st.write("• Ensure the Export sheet formulas reference the correct Prices sheet cells")
                                st.write("• Check that formulas use the format: `=Prices!D4`, `=Prices!D5`, etc.")
                                st.write("• Verify the Prices sheet is named exactly 'Prices'")
                        
                        if len(preview_df) > 10:
                            st.write(f"... and {len(preview_df) - 10} more rows")
                        os.unlink(tmp.name)
                except Exception as e:
                    st.error(f"Error previewing: {e}")
            
            # Download button
            try:
                csv_data = export_sheet_to_csv(workbook, 'Export')
                if csv_data:
                    st.download_button(
                        label="📥 Download Export Sheet as CSV",
                        data=csv_data,
                        file_name="export_data.csv",
                        mime="text/csv",
                        type="primary"
                    )
                else:
                    st.error("❌ Failed to export CSV")
            except Exception as e:
                st.error(f"❌ Export error: {e}")
        else:
            st.warning("⚠️ 'Export' sheet not found")
            st.write("Available sheets:")
            for sheet in workbook.sheetnames:
                st.write(f"• {sheet}")
    
    # Display current file info and diagnostics
    st.markdown("---")
    with st.expander("ℹ️ File Information & Diagnostics"):
        st.write(f"**File Path:** {excel_file_path if not temp_file_path else 'Uploaded File'}")
        st.write(f"**Available Sheets:** {', '.join(workbook.sheetnames)}")
        st.write(f"**Target Cells:** D4, D5, D6, D7 in 'Prices' sheet")
        st.write(f"**Export Sheet:** 'Export' sheet → CSV download")
        
        # Diagnostic: Check formulas in Export sheet
        if 'Export' in workbook.sheetnames:
            st.subheader("🔍 Formula Diagnostics")
            export_sheet = workbook['Export']
            
            # Check for formulas that might reference Prices sheet
            formulas_found = []
            for row in export_sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula = str(cell.value)
                        if 'Prices!' in formula or any(f'D{i}' in formula for i in [4,5,6,7]):
                            formulas_found.append({
                                'Cell': cell.coordinate,
                                'Formula': formula,
                                'Current Value': cell.value
                            })
            
            if formulas_found:
                st.write("**Formulas referencing Prices sheet:**")
                for formula_info in formulas_found[:10]:  # Show first 10
                    st.write(f"• {formula_info['Cell']}: `{formula_info['Formula']}`")
                if len(formulas_found) > 10:
                    st.write(f"... and {len(formulas_found) - 10} more formulas")
            else:
                st.warning("⚠️ No formulas found that reference the Prices sheet or cells D4-D7")
                st.write("**This might explain why Variant Price is blank.**")
                st.write("**Expected formula examples:**")
                st.write("• `=Prices!D4` (direct reference)")
                st.write("• `=IF(A2=\"Product1\",Prices!D4,Prices!D5)` (conditional)")
                st.write("• `=VLOOKUP(B2,Prices!D4:D7,1,FALSE)` (lookup)")
        
        # Show current values in Prices sheet
        if 'Prices' in workbook.sheetnames:
            current_prices = get_current_prices(workbook)
            st.subheader("📊 Current Prices Sheet Values")
            for cell, value in current_prices.items():
                st.write(f"• {cell}: `{value}`")
    
    # Clean up
    if workbook:
        workbook.close()
    
    # Clean up temp file
    if temp_file_path and os.path.exists(temp_file_path):
        try:
            os.unlink(temp_file_path)
        except:
            pass

if __name__ == "__main__":
    main()
